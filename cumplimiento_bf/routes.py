# -*- coding: utf-8 -*-
"""
Rutas del módulo de Cumplimiento BF
Maneja las rutas y endpoints relacionados con el seguimiento de SKUs de Black Friday
"""

from flask import request, render_template, jsonify, send_file
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO

from cumplimiento_bf.blueprint import bp
from config import MAZATLAN_TZ as mazatlan_tz
from database import get_fresh_data, obtener_mes_actual
from cumplimiento_bf.services import (
    calcular_cumplimiento_skus,
    obtener_categorias_bf,
    obtener_inventario_ventas_bf,
    agrupar_inventario_por_tipo,
    agrupar_inventario_por_tipo_desde_skus
)


# Canales permitidos para BF
CANALES_BF = ['Mercado Libre', 'CrediTienda', 'Walmart', 'Shein', 'Yuhu', 'Liverpool', 'AliExpress', 'Aliexpress']


@bp.route("/cumplimiento-bf", methods=["GET", "POST"])
def cumplimiento_bf():
    """Página principal de cumplimiento de metas BF"""

    # Inicializar variables por defecto para evitar errores
    skus_data = []
    resumen_general = {
        'total_skus': 0,
        'total_real': 0,
        'total_cantidad': 0,
        'total_ingreso_real': 0,
        'total_costo': 0,
        'total_gastos_directos': 0,
        'roi_promedio': 0
    }
    inventario_agrupado = {'resumen': [], 'productos_por_tipo': {}}
    categorias_bf = []
    periodo_texto = "Mes completo de noviembre"
    error = None

    try:
        # Obtener mes actual
        mes_actual = obtener_mes_actual()
        mes_seleccionado = mes_actual

        # Cargar datos frescos
        print(f"INFO: Cargando datos frescos para Cumplimiento BF (mes {mes_seleccionado})...")
        df_ventas, _, _ = get_fresh_data(mes_seleccionado)

        # Obtener categorías disponibles desde el catálogo BF
        categorias_bf = obtener_categorias_bf()

        # Verificar si hay datos de ventas
        if df_ventas.empty:
            error = "No hay datos de ventas disponibles para el mes actual"
            print(f"WARNING: {error}")
        else:
            # Obtener parámetros de filtros
            hoy = datetime.now(mazatlan_tz).replace(hour=0, minute=0, second=0, microsecond=0)

            if request.method == "POST":
                preset = request.form.get("preset", "mes_completo")
                filtro_tipo = request.form.get("filtro_tipo", "todos")
                filtro_canal = request.form.get("filtro_canal", "todos")
                filtro_categoria = request.form.get("filtro_categoria", "todas")
                rango_personalizado = request.form.get("rango_personalizado", "")
            else:  # GET - valores por defecto (mes completo)
                preset = "mes_completo"
                filtro_tipo = "todos"
                filtro_canal = "todos"
                filtro_categoria = "todas"
                rango_personalizado = ""

            # Guardar valores originales para el template
            filtro_tipo_original = filtro_tipo
            filtro_canal_original = filtro_canal
            filtro_categoria_original = filtro_categoria

            # Convertir "todos" a None para los servicios
            filtro_tipo = None if filtro_tipo == "todos" else filtro_tipo
            filtro_canal = None if filtro_canal == "todos" else filtro_canal
            filtro_categoria = None if filtro_categoria == "todas" else filtro_categoria

            # Determinar fechas según el preset
            if preset == "hoy":
                f1 = hoy
                f2 = hoy + timedelta(days=1)
                periodo_texto = f"Hoy ({hoy.strftime('%d/%m/%Y')})"
                print(f"INFO: Período seleccionado - Hoy: {hoy.strftime('%Y-%m-%d')}")
            elif preset == "7":
                f1 = hoy - timedelta(days=7)
                f2 = hoy + timedelta(days=1)
                periodo_texto = "Últimos 7 días"
                print(f"INFO: Período seleccionado - Últimos 7 días: {f1.strftime('%Y-%m-%d')} a {(f2 - timedelta(days=1)).strftime('%Y-%m-%d')}")
            elif preset == "mes_completo":
                # Mes actual completo (del 1 al último día del mes)
                f1 = hoy.replace(day=1)
                ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                f2 = ultimo_dia + timedelta(days=1)
                mes_nombre = f1.strftime('%B')
                periodo_texto = f"Mes completo de {mes_nombre}"
                print(f"INFO: Período seleccionado - Mes completo: {f1.strftime('%Y-%m-%d')} a {ultimo_dia.strftime('%Y-%m-%d')}")
            elif preset == "personalizado":
                # Validar que exista el rango personalizado
                print(f"DEBUG: Rango personalizado recibido: '{rango_personalizado}' (tipo: {type(rango_personalizado)}, longitud: {len(rango_personalizado)})")

                if not rango_personalizado or rango_personalizado.strip() == "":
                    error = "Por favor selecciona un rango de fechas personalizado"
                    print(f"WARNING: {error}")
                    # Fallback a mes completo
                    preset = "mes_completo"
                    f1 = hoy.replace(day=1)
                    ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                    f2 = ultimo_dia + timedelta(days=1)
                    mes_nombre = f1.strftime('%B')
                    periodo_texto = f"Mes completo de {mes_nombre}"
                else:
                    # Limpiar el string y buscar separador
                    rango_limpio = rango_personalizado.strip()
                    print(f"DEBUG: Rango limpio: '{rango_limpio}'")
                    print(f"DEBUG: Bytes del rango: {rango_limpio.encode('utf-8')}")

                    # Intentar varios separadores posibles (orden de prioridad)
                    separador = None
                    if " a " in rango_limpio:
                        separador = " a "
                    elif " to " in rango_limpio:
                        separador = " to "
                    elif " al " in rango_limpio.lower():
                        separador = " al "
                    elif " - " in rango_limpio:
                        separador = " - "
                    elif "to" in rango_limpio:
                        separador = "to"
                    elif "a" in rango_limpio and rango_limpio.count("a") == 1:
                        # Si solo hay una "a" podría ser el separador sin espacios
                        separador = "a"

                    # Caso 1: Si hay separador, es un rango
                    if separador:
                        try:
                            partes = rango_limpio.split(separador)
                            print(f"DEBUG: Separador detectado: '{separador}', Partes: {partes}")

                            if len(partes) == 2:
                                f1_str = partes[0].strip()
                                f2_str = partes[1].strip()
                                print(f"DEBUG: Parseando fechas - Inicio: '{f1_str}', Fin: '{f2_str}'")

                                f1 = mazatlan_tz.localize(datetime.strptime(f1_str, "%Y-%m-%d"))
                                f2_temp = mazatlan_tz.localize(datetime.strptime(f2_str, "%Y-%m-%d"))
                                f2 = f2_temp + timedelta(days=1)
                                periodo_texto = f"Personalizado ({f1.strftime('%d/%m/%Y')} - {f2_temp.strftime('%d/%m/%Y')})"
                                print(f"OK: Período personalizado configurado: {f1.strftime('%Y-%m-%d')} a {f2_temp.strftime('%Y-%m-%d')}")
                            else:
                                raise ValueError(f"Se esperaban 2 partes pero se obtuvieron {len(partes)}")
                        except ValueError as e:
                            error = f"Formato de fecha inválido. Por favor usa el selector de fechas. Error: {str(e)}"
                            print(f"ERROR: {error}")
                            # Fallback a mes completo
                            preset = "mes_completo"
                            f1 = hoy.replace(day=1)
                            ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                            f2 = ultimo_dia + timedelta(days=1)
                            mes_nombre = f1.strftime('%B')
                            periodo_texto = f"Mes completo de {mes_nombre}"
                    # Caso 2: No hay separador, es un solo día en formato YYYY-MM-DD
                    else:
                        try:
                            print(f"DEBUG: No se encontró separador, asumiendo un solo día: '{rango_limpio}'")
                            # Parsear como un solo día
                            f1 = mazatlan_tz.localize(datetime.strptime(rango_limpio, "%Y-%m-%d"))
                            # Para un solo día, f2 es el día siguiente
                            f2 = f1 + timedelta(days=1)
                            periodo_texto = f"Personalizado ({f1.strftime('%d/%m/%Y')})"
                            print(f"OK: Día personalizado configurado: {f1.strftime('%Y-%m-%d')}")
                        except ValueError as e:
                            error = f"Formato de fecha inválido. Por favor usa el selector de fechas. Error: {str(e)}"
                            print(f"ERROR: {error}")
                            # Fallback a mes completo
                            preset = "mes_completo"
                            f1 = hoy.replace(day=1)
                            ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                            f2 = ultimo_dia + timedelta(days=1)
                            mes_nombre = f1.strftime('%B')
                            periodo_texto = f"Mes completo de {mes_nombre}"
            else:
                # Default a mes completo
                f1 = hoy.replace(day=1)
                ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                f2 = ultimo_dia + timedelta(days=1)
                mes_nombre = f1.strftime('%B')
                periodo_texto = f"Mes completo de {mes_nombre}"
                print(f"INFO: Período por defecto - Mes completo: {f1.strftime('%Y-%m-%d')} a {ultimo_dia.strftime('%Y-%m-%d')}")

            print(f"INFO: Calculando cumplimiento BF para período {f1} - {f2}")

            # Calcular cumplimiento de SKUs BF (sin metas por ahora)
            skus_data, resumen_general = calcular_cumplimiento_skus(
                df_ventas, pd.DataFrame(), f1, f2, filtro_tipo, filtro_canal, filtro_categoria
            )

            # Obtener datos de inventario agrupados por tipo (usando get_ventas_producto_compra_periodo con fechas)
            # Las existencias son generales, pero las ventas se filtran por canal
            inventario_agrupado = agrupar_inventario_por_tipo(filtro_tipo, filtro_categoria, filtro_canal, f1, f2)

            print(f"OK: Cumplimiento BF calculado - {len(skus_data)} SKUs, ${resumen_general['total_real']:,.0f} en ventas")

    except Exception as e:
        error = f"Error procesando datos: {str(e)}"
        print(f"ERROR en cumplimiento_bf: {e}")
        import traceback
        traceback.print_exc()
        filtro_tipo_original = None
        filtro_canal_original = None
        filtro_categoria_original = None
        mes_seleccionado = obtener_mes_actual()

    # Renderizar template con todas las variables necesarias
    return render_template("cumplimiento_bf.html",
                         error=error,
                         skus_data=skus_data,
                         resumen_general=resumen_general,
                         inventario_agrupado=inventario_agrupado,
                         canales_bf=CANALES_BF,
                         categorias_bf=categorias_bf,
                         filtro_tipo=filtro_tipo_original if 'filtro_tipo_original' in locals() else None,
                         filtro_canal=filtro_canal_original if 'filtro_canal_original' in locals() else None,
                         filtro_categoria=filtro_categoria_original if 'filtro_categoria_original' in locals() else None,
                         periodo_texto=periodo_texto,
                         preset_actual=preset if 'preset' in locals() else "mes_completo",
                         rango_personalizado_valor=rango_personalizado if 'rango_personalizado' in locals() else "",
                         mes_seleccionado=mes_seleccionado if 'mes_seleccionado' in locals() else obtener_mes_actual())


@bp.route("/cumplimiento-bf-ajax", methods=["POST"])
def cumplimiento_bf_ajax():
    """Endpoint AJAX para actualizar datos de cumplimiento BF"""
    try:
        # Cargar datos frescos
        mes_actual = obtener_mes_actual()
        df_ventas, _, _ = get_fresh_data(mes_actual)

        if df_ventas.empty:
            return jsonify({
                'success': False,
                'error': 'No se encontraron datos para el filtro aplicado'
            })

        # Obtener parámetros del formulario AJAX
        hoy = datetime.now(mazatlan_tz).replace(hour=0, minute=0, second=0, microsecond=0)
        preset = request.form.get("preset", "7")
        filtro_tipo = request.form.get("filtro_tipo", "todos")
        filtro_canal = request.form.get("filtro_canal", "todos")
        filtro_categoria = request.form.get("filtro_categoria", "todas")
        rango_personalizado = request.form.get("rango_personalizado", "")

        # Convertir "todos" a None para los servicios
        filtro_tipo = None if filtro_tipo == "todos" else filtro_tipo
        filtro_canal = None if filtro_canal == "todos" else filtro_canal
        filtro_categoria = None if filtro_categoria == "todas" else filtro_categoria

        # Determinar fechas según el preset
        if preset == "hoy":
            f1 = hoy
            f2 = hoy + timedelta(days=1)
            print(f"AJAX: Período - Hoy: {hoy.strftime('%Y-%m-%d')}")
        elif preset == "7":
            f1 = hoy - timedelta(days=7)
            f2 = hoy + timedelta(days=1)
            print(f"AJAX: Período - Últimos 7 días: {f1.strftime('%Y-%m-%d')} a {(f2 - timedelta(days=1)).strftime('%Y-%m-%d')}")
        elif preset == "mes_completo":
            # Mes actual completo (del 1 al último día del mes)
            f1 = hoy.replace(day=1)
            ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            f2 = ultimo_dia + timedelta(days=1)
            print(f"AJAX: Período - Mes completo: {f1.strftime('%Y-%m-%d')} a {ultimo_dia.strftime('%Y-%m-%d')}")
        elif preset == "personalizado":
            print(f"AJAX: Rango personalizado recibido: '{rango_personalizado}'")

            if not rango_personalizado or rango_personalizado.strip() == "":
                return jsonify({
                    'success': False,
                    'error': 'Por favor selecciona un rango de fechas personalizado'
                })

            # Limpiar el string y buscar separador
            rango_limpio = rango_personalizado.strip()

            # Intentar varios separadores posibles (orden de prioridad)
            separador = None
            if " a " in rango_limpio:
                separador = " a "
            elif " to " in rango_limpio:
                separador = " to "
            elif " al " in rango_limpio.lower():
                separador = " al "
            elif " - " in rango_limpio:
                separador = " - "
            elif "to" in rango_limpio:
                separador = "to"
            elif "a" in rango_limpio and rango_limpio.count("a") == 1:
                # Si solo hay una "a" podría ser el separador sin espacios
                separador = "a"

            if separador:
                try:
                    partes = rango_limpio.split(separador)
                    if len(partes) == 2:
                        f1_str = partes[0].strip()
                        f2_str = partes[1].strip()
                        print(f"AJAX: Parseando fechas - Inicio: '{f1_str}', Fin: '{f2_str}'")

                        f1 = mazatlan_tz.localize(datetime.strptime(f1_str, "%Y-%m-%d"))
                        f2_temp = mazatlan_tz.localize(datetime.strptime(f2_str, "%Y-%m-%d"))
                        f2 = f2_temp + timedelta(days=1)
                        print(f"AJAX: Período personalizado: {f1.strftime('%Y-%m-%d')} a {f2_temp.strftime('%Y-%m-%d')}")
                    else:
                        return jsonify({
                            'success': False,
                            'error': f'Formato inválido: se esperaban 2 fechas'
                        })
                except ValueError as e:
                    return jsonify({
                        'success': False,
                        'error': f'Formato de fecha inválido: {str(e)}'
                    })
            else:
                return jsonify({
                    'success': False,
                    'error': f'No se encontró separador válido en el rango'
                })
        else:
            # Default a mes completo
            f1 = hoy.replace(day=1)
            ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            f2 = ultimo_dia + timedelta(days=1)
            print(f"AJAX: Período por defecto - Mes completo: {f1.strftime('%Y-%m-%d')} a {ultimo_dia.strftime('%Y-%m-%d')}")

        # Calcular cumplimiento
        skus_data, resumen_general = calcular_cumplimiento_skus(
            df_ventas, pd.DataFrame(), f1, f2, filtro_tipo, filtro_canal, filtro_categoria
        )

        return jsonify({
            'success': True,
            'skus_data': skus_data,
            'resumen_general': resumen_general
        })

    except Exception as e:
        print(f"Error en AJAX cumplimiento BF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Error procesando datos: {str(e)}'
        })


@bp.route("/cumplimiento-bf-exportar", methods=["POST"])
def cumplimiento_bf_exportar():
    """Endpoint para exportar datos de cumplimiento BF a Excel (2 hojas con formato profesional)"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Cargar datos frescos
        mes_actual = obtener_mes_actual()
        df_ventas, _, _ = get_fresh_data(mes_actual)

        if df_ventas.empty:
            return jsonify({'success': False, 'error': 'No hay datos disponibles para exportar'})

        # Obtener parámetros
        hoy = datetime.now(mazatlan_tz).replace(hour=0, minute=0, second=0, microsecond=0)
        preset = request.form.get("preset", "mes_completo")
        filtro_tipo = request.form.get("filtro_tipo", "todos")
        filtro_canal = request.form.get("filtro_canal", "todos")
        filtro_categoria = request.form.get("filtro_categoria", "todas")
        periodo_texto = request.form.get("periodo_texto", "Mes completo de noviembre")

        # Convertir "todos" a None
        filtro_tipo_param = None if filtro_tipo == "todos" else filtro_tipo
        filtro_canal_param = None if filtro_canal == "todos" else filtro_canal
        filtro_categoria_param = None if filtro_categoria == "todas" else filtro_categoria

        print(f"=== EXPORTAR CON FILTROS ===")
        print(f"Preset: {preset}")
        print(f"Filtro Tipo: {filtro_tipo} -> {filtro_tipo_param}")
        print(f"Filtro Canal: {filtro_canal} -> {filtro_canal_param}")
        print(f"Filtro Categoría: {filtro_categoria} -> {filtro_categoria_param}")

        # Determinar fechas
        if preset == "hoy":
            f1 = hoy
            f2 = hoy + timedelta(days=1)
        elif preset == "7":
            f1 = hoy - timedelta(days=7)
            f2 = hoy + timedelta(days=1)
        elif preset == "mes_completo":
            f1 = hoy.replace(day=1)
            ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            f2 = ultimo_dia + timedelta(days=1)
        elif preset == "personalizado":
            rango = request.form.get("rango_personalizado", "")
            if rango:
                rango_limpio = rango.strip()
                # Intentar varios separadores posibles
                separador = None
                if " a " in rango_limpio:
                    separador = " a "
                elif " to " in rango_limpio:
                    separador = " to "
                elif " al " in rango_limpio.lower():
                    separador = " al "
                elif " - " in rango_limpio:
                    separador = " - "
                elif "to" in rango_limpio:
                    separador = "to"

                # Si hay separador, es un rango
                if separador:
                    f1_str, f2_str = rango_limpio.split(separador)
                    f1 = mazatlan_tz.localize(datetime.strptime(f1_str.strip(), "%Y-%m-%d"))
                    f2_temp = mazatlan_tz.localize(datetime.strptime(f2_str.strip(), "%Y-%m-%d"))
                    f2 = f2_temp + timedelta(days=1)
                else:
                    # No hay separador, es un solo día
                    f1 = mazatlan_tz.localize(datetime.strptime(rango_limpio, "%Y-%m-%d"))
                    f2 = f1 + timedelta(days=1)
            else:
                # Fallback a mes completo
                f1 = hoy.replace(day=1)
                ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                f2 = ultimo_dia + timedelta(days=1)
        else:
            f1 = hoy.replace(day=1)
            ultimo_dia = (hoy.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            f2 = ultimo_dia + timedelta(days=1)

        # Calcular cumplimiento (para Hoja 2)
        skus_data, resumen_general = calcular_cumplimiento_skus(
            df_ventas, pd.DataFrame(), f1, f2, filtro_tipo_param, filtro_canal_param, filtro_categoria_param
        )

        # Obtener inventario (para Hoja 1) con ventas del período usando arrayJoin
        # Las existencias son generales, pero las ventas se filtran por canal
        inventario_data = obtener_inventario_ventas_bf(filtro_tipo_param, filtro_categoria_param, filtro_canal_param, f1, f2)

        # Crear archivo Excel en memoria
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # ========================================
            # HOJA 1: INVENTARIO Y VENTAS DEL MES
            # ========================================
            if inventario_data:
                df_inv = pd.DataFrame(inventario_data)

                # Preparar datos con tipo - las ventas ya vienen del período correcto en Venta_Mes_Actual
                df_inv_export = pd.DataFrame({
                    'Tipo': df_inv.apply(lambda x: 'Relevante' if x['Es_Relevante'] else ('Nuevo' if x['Es_Nuevo'] else 'Remate'), axis=1),
                    'SKU': df_inv['sku'],
                    'Descripción': df_inv['descripcion'],
                    'Categoría': df_inv['categoria'],
                    'Existencia': pd.to_numeric(df_inv['Existencia'], errors='coerce').fillna(0).astype(int),
                    'Venta Mes': pd.to_numeric(df_inv['Venta_Mes_Actual'], errors='coerce').fillna(0).astype(int)
                })

                # Ordenar por Tipo y SKU
                orden_tipo = {'Relevante': 1, 'Nuevo': 2, 'Remate': 3}
                df_inv_export['orden'] = df_inv_export['Tipo'].map(orden_tipo)
                df_inv_export = df_inv_export.sort_values(['orden', 'SKU']).drop('orden', axis=1)

                # Escribir a Excel empezando en fila 5 (dejando espacio para título)
                df_inv_export.to_excel(writer, sheet_name='Inventario y Ventas', index=False, startrow=4)

                # Obtener worksheet para aplicar formato
                ws1 = writer.sheets['Inventario y Ventas']

                # Título principal (fila 2)
                ws1.merge_cells('A2:F2')
                titulo_cell = ws1['A2']
                titulo_cell.value = 'INVENTARIO Y VENTAS DEL MES - CUMPLIMIENTO BF (DESGLOSE DETALLADO)'
                titulo_cell.font = Font(size=14, bold=True, color='FFFFFF')
                titulo_cell.fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
                titulo_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Subtítulo con filtros aplicados (fila 3)
                ws1.merge_cells('A3:F3')
                subtitulo_cell = ws1['A3']
                filtros_texto = []
                if filtro_tipo_param:
                    filtros_texto.append(f'Tipo: {filtro_tipo_param}')
                if filtro_canal_param:
                    filtros_texto.append(f'Canal: {filtro_canal_param}')
                if filtro_categoria_param:
                    filtros_texto.append(f'Categoría: {filtro_categoria_param}')
                filtros_str = ' | '.join(filtros_texto) if filtros_texto else 'Sin filtros'
                subtitulo_cell.value = f'Período: {periodo_texto} | Filtros: {filtros_str}'
                subtitulo_cell.font = Font(size=10, italic=True)
                subtitulo_cell.alignment = Alignment(horizontal='left', vertical='center')

                # Formato de headers (fila 5)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col in range(1, 7):
                    cell = ws1.cell(row=5, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Formato de datos y colores por tipo
                colores_tipo = {
                    'Relevante': 'E9D5FF',  # Morado claro
                    'Nuevo': 'DBEAFE',      # Azul claro
                    'Remate': 'FED7AA'      # Naranja claro
                }

                for row_idx, row in enumerate(df_inv_export.itertuples(), start=6):
                    tipo_valor = row.Tipo
                    color_fondo = colores_tipo.get(tipo_valor, 'FFFFFF')

                    for col_idx in range(1, 7):
                        cell = ws1.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type='solid')
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left' if col_idx in [1, 2, 3, 4] else 'right', vertical='center')

                        # Formato de números
                        if col_idx == 5:  # Existencia
                            cell.number_format = '#,##0'
                        elif col_idx == 6:  # Venta Mes
                            cell.number_format = '#,##0'

                # Ajustar ancho de columnas
                ws1.column_dimensions['A'].width = 12  # Tipo
                ws1.column_dimensions['B'].width = 12  # SKU
                ws1.column_dimensions['C'].width = 50  # Descripción
                ws1.column_dimensions['D'].width = 20  # Categoría
                ws1.column_dimensions['E'].width = 12  # Existencia
                ws1.column_dimensions['F'].width = 12  # Venta Mes

                # Ajustar altura de título
                ws1.row_dimensions[2].height = 25

            # ========================================
            # HOJA 2: DETALLE POR SKU (CON DESGLOSE INDIVIDUAL/COMBO)
            # ========================================
            if skus_data:
                df_skus = pd.DataFrame(skus_data)

                # Preparar datos con identificador de tipo de fila
                df_skus_export = pd.DataFrame({
                    'SKU': df_skus.apply(lambda x: x['sku'] if x.get('tipo_fila') == 'individual' else '', axis=1),
                    'Tipo': df_skus.apply(lambda x: 'Combo' if x.get('tipo_fila') == 'combo' else ('Relevante' if x['Es_Relevante'] else ('Nuevo' if x['Es_Nuevo'] else 'Remate')), axis=1),
                    'Descripción': df_skus['descripcion'],
                    'Categoría': df_skus.apply(lambda x: x['categoria'] if x.get('tipo_fila') == 'individual' else '', axis=1),
                    'Cantidad': df_skus['Cantidad_Vendida'],
                    'Ventas': df_skus['Ventas_Reales'],
                    'Costo Venta': df_skus['Costo_Venta'],
                    'Gastos Directos': df_skus['Gastos_Directos'],
                    'Ingreso Real': df_skus['Ingreso_Real'],
                    'ROI %': df_skus['ROI'],
                    'tipo_fila': df_skus.get('tipo_fila', 'individual')  # Mantener para formato
                })

                # Escribir a Excel
                df_skus_export[['SKU', 'Tipo', 'Descripción', 'Categoría', 'Cantidad', 'Ventas', 'Costo Venta', 'Gastos Directos', 'Ingreso Real', 'ROI %']].to_excel(
                    writer, sheet_name='Detalle por SKU', index=False, startrow=4
                )

                # Obtener worksheet
                ws2 = writer.sheets['Detalle por SKU']

                # Título principal (fila 2)
                ws2.merge_cells('A2:J2')
                titulo_cell = ws2['A2']
                titulo_cell.value = 'DETALLE POR SKU - CUMPLIMIENTO BF (Individual + Combo)'
                titulo_cell.font = Font(size=14, bold=True, color='FFFFFF')
                titulo_cell.fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
                titulo_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Subtítulo con filtros aplicados (fila 3)
                ws2.merge_cells('A3:J3')
                subtitulo_cell = ws2['A3']
                filtros_texto = []
                if filtro_tipo_param:
                    filtros_texto.append(f'Tipo: {filtro_tipo_param}')
                if filtro_canal_param:
                    filtros_texto.append(f'Canal: {filtro_canal_param}')
                if filtro_categoria_param:
                    filtros_texto.append(f'Categoría: {filtro_categoria_param}')
                filtros_str = ' | '.join(filtros_texto) if filtros_texto else 'Sin filtros'
                subtitulo_cell.value = f'Período: {periodo_texto} | Filtros: {filtros_str}'
                subtitulo_cell.font = Font(size=10, italic=True)
                subtitulo_cell.alignment = Alignment(horizontal='left', vertical='center')

                # Formato de headers
                for col in range(1, 11):
                    cell = ws2.cell(row=5, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Formato de datos
                for row_idx, (idx, row) in enumerate(df_skus_export.iterrows(), start=6):
                    tipo_fila = row.get('tipo_fila', 'individual')
                    tipo_valor = row['Tipo']

                    # Color de fondo según tipo de fila
                    if tipo_fila == 'combo':
                        color_fondo = 'F8F9FA'  # Gris claro para combos
                    else:
                        color_fondo = colores_tipo.get(tipo_valor, 'FFFFFF')

                    for col_idx in range(1, 11):
                        cell = ws2.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type='solid')
                        cell.border = border

                        # Alineación especial para filas combo
                        if tipo_fila == 'combo' and col_idx == 3:  # Columna Descripción
                            # Indentar texto para filas combo
                            cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                            cell.font = Font(italic=True, size=10)
                        else:
                            cell.alignment = Alignment(horizontal='left' if col_idx in [1, 2, 3, 4] else 'right', vertical='center')

                        # Color de fuente más tenue para combos
                        if tipo_fila == 'combo':
                            cell.font = Font(italic=True, color='6C757D', size=10)

                        # Formato de números
                        if col_idx == 5:  # Cantidad
                            cell.number_format = '#,##0'
                        elif col_idx in [6, 7, 8, 9]:  # Ventas, Costo, Gastos, Ingreso
                            cell.number_format = '$#,##0'
                        elif col_idx == 10:  # ROI %
                            cell.number_format = '0.00"%"'

                # Fila de TOTALES
                ultima_fila = len(df_skus_export) + 6

                # Calcular totales SOLO de filas principales (individual)
                df_solo_individual = df_skus_export[df_skus_export['tipo_fila'] == 'individual']
                total_cantidad = df_solo_individual['Cantidad'].sum()
                total_ventas = df_solo_individual['Ventas'].sum()
                total_costo = df_solo_individual['Costo Venta'].sum()
                total_gastos = df_solo_individual['Gastos Directos'].sum()
                total_ingreso = df_solo_individual['Ingreso Real'].sum()
                total_roi = (total_ingreso / total_costo * 100) if total_costo > 0 else 0

                # Aplicar totales
                ws2.merge_cells(f'A{ultima_fila}:D{ultima_fila}')
                total_cell = ws2[f'A{ultima_fila}']
                total_cell.value = 'TOTALES (Solo Individual)'
                total_cell.font = Font(bold=True, size=11)
                total_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                total_cell.alignment = Alignment(horizontal='center', vertical='center')
                total_cell.border = border

                # Valores de totales
                totales_valores = [total_cantidad, total_ventas, total_costo, total_gastos, total_ingreso, total_roi]
                for col_idx, valor in enumerate(totales_valores, start=5):
                    cell = ws2.cell(row=ultima_fila, column=col_idx)
                    cell.value = valor
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border

                    if col_idx == 5:
                        cell.number_format = '#,##0'
                    elif col_idx in [6, 7, 8, 9]:
                        cell.number_format = '$#,##0'
                    elif col_idx == 10:
                        cell.number_format = '0.00"%"'

                # Ajustar ancho de columnas
                ws2.column_dimensions['A'].width = 12
                ws2.column_dimensions['B'].width = 12
                ws2.column_dimensions['C'].width = 50
                ws2.column_dimensions['D'].width = 20
                ws2.column_dimensions['E'].width = 10
                ws2.column_dimensions['F'].width = 12
                ws2.column_dimensions['G'].width = 12
                ws2.column_dimensions['H'].width = 14
                ws2.column_dimensions['I'].width = 12
                ws2.column_dimensions['J'].width = 10

                # Ajustar altura de título
                ws2.row_dimensions[2].height = 25

        output.seek(0)

        # Nombre del archivo
        fecha_str = datetime.now().strftime('%Y%m%d')
        filename = f"Cumplimiento_BF_{fecha_str}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error exportando a Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Error exportando: {str(e)}'
        })
